#!/usr/bin/env python3
"""
Test de verificación: herramientas de creación de documentos.

Verifica docx_create, xlsx_create, pptx_create y pdf_create
con datos reales. Los archivos se generan en tests/output/.

Uso:
    python tests/test_documents.py
"""
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from mcp_windows_server import docx_create, xlsx_create, pptx_create, pdf_create

OUT = Path(__file__).parent / "output"
OUT.mkdir(exist_ok=True)

failures = 0


def check(name: str, result_json: str) -> None:
    global failures
    data = json.loads(result_json)
    if "error" in data:
        print(f"  ❌ {name}: {data['error']}")
        failures += 1
    else:
        print(f"  ✅ {name}: {data.get('path')}")


print("1/4 Word...")
check("docx_create", docx_create(
    path=str(OUT / "test.docx"),
    content=json.dumps({
        "title": "Informe de Prueba",
        "sections": [
            {"heading": "Resumen", "paragraphs": ["Texto de prueba."],
             "bullets": ["Punto A", "Punto B"]},
            {"heading": "Tabla", "table": {"headers": ["A", "B"],
             "rows": [["1", "2"], ["3", "4"]]}},
        ],
    }),
))

print("2/4 Excel con fórmulas...")
check("xlsx_create", xlsx_create(
    path=str(OUT / "test.xlsx"),
    content=json.dumps({
        "sheets": [{
            "name": "Datos",
            "headers": ["Item", "Precio", "Cantidad", "Total"],
            "rows": [["A", 10, 2, "=B2*C2"], ["B", 20, 3, "=B3*C3"]],
            "formulas": {"D5": "=SUM(D2:D3)", "D6": "=AVERAGE(D2:D3)"},
        }],
    }),
))

print("3/4 PowerPoint...")
check("pptx_create", pptx_create(
    path=str(OUT / "test.pptx"),
    content=json.dumps({
        "title": "Test", "subtitle": "Automatizado",
        "slides": [{"title": "S1", "bullets": ["a", "b"]}],
    }),
))

print("4/4 PDF con membrete (incluye bullets consecutivos - regresión fpdf2)...")
check("pdf_create", pdf_create(
    path=str(OUT / "test.pdf"),
    content=json.dumps({
        "letterhead": {"company": "Test S.A.", "color": "#1a5276",
                       "footer": "Pie de prueba"},
        "title": "Doc de Prueba",
        "sections": [
            {"heading": "Sección", "text": "Contenido."},
            {"bullets": ["Bullet 1", "Bullet 2", "Bullet 3"]},
            {"table": {"headers": ["X", "Y"], "rows": [["1", "2"]]}},
        ],
    }),
))

# Verificación de fórmulas Excel (regresión: openpyxl debe guardarlas como fórmulas)
try:
    import openpyxl
    wb = openpyxl.load_workbook(OUT / "test.xlsx")
    ws = wb.active
    assert ws["D5"].value == "=SUM(D2:D3)", f"D5 incorrecto: {ws['D5'].value}"
    assert ws["D2"].value == "=B2*C2", f"D2 incorrecto: {ws['D2'].value}"
    print("  ✅ Fórmulas Excel verificadas en el archivo")
except Exception as e:
    print(f"  ❌ Verificación de fórmulas: {e}")
    failures += 1

print()
if failures:
    print(f"❌ {failures} FALLOS")
    sys.exit(1)
print("✅ TODOS LOS TESTS PASARON")
